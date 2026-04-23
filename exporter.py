import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COR_HEADER_BG  = "1B6B5A"
COR_HEADER_FG  = "FFFFFF"
COR_ENTRADA_BG = "E6F4F1"
COR_SAIDA_BG   = "FDE8E8"
COR_TOTAL_BG   = "0D3B2E"
COR_BORDA      = "C5D8D4"

# Formato correto para moeda BR no Excel: texto "R$" entre aspas + número
FMT_MOEDA = '"R$" #,##0.00'
FMT_DATA  = 'DD/MM/YYYY'

COLUNAS = [
    ("Data",               13),
    ("Descrição",          52),
    ("Tipo",               11),
    ("Valor (R$)",         16),
    ("Saldo (R$)",         16),
    ("Banco / Origem",     22),
    ("Forma de Pagamento", 24),
    ("Categoria",          18),
    ("Apuração",           15),
    ("Observação",         25),
]


def _borda():
    lado = Side(style="thin", color=COR_BORDA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _cell(ws, row, col, value, fill, font, align, borda, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill
    c.font   = font
    c.alignment = align
    c.border = borda
    if fmt:
        c.number_format = fmt
    return c


def generate_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação"

    borda       = _borda()
    fill_header = PatternFill("solid", fgColor=COR_HEADER_BG)
    fill_total  = PatternFill("solid", fgColor=COR_TOTAL_BG)
    fill_ent    = PatternFill("solid", fgColor=COR_ENTRADA_BG)
    fill_sai    = PatternFill("solid", fgColor=COR_SAIDA_BG)
    font_header = Font(bold=True, color=COR_HEADER_FG, name="Calibri", size=11)
    font_total  = Font(bold=True, color=COR_HEADER_FG, name="Calibri", size=11)
    font_data   = Font(name="Calibri", size=10)
    al_center   = Alignment(horizontal="center", vertical="center")
    al_left     = Alignment(horizontal="left",   vertical="center")
    al_right    = Alignment(horizontal="right",  vertical="center")

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    for col_idx, (nome, largura) in enumerate(COLUNAS, start=1):
        _cell(ws, 1, col_idx, nome, fill_header, font_header, al_center, borda)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.row_dimensions[1].height = 24

    # ── Dados ────────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = fill_ent if row.tipo == "Entrada" else fill_sai

        # Data como objeto datetime real (não string) → Excel renderiza como data
        data_val = row.data.to_pydatetime() if pd.notna(row.data) else None
        valor    = float(row.valor) if pd.notna(row.valor) else None
        saldo    = float(row.saldo) if pd.notna(row.saldo) else None

        linha = [
            (data_val,         al_center, FMT_DATA),
            (row.descricao,    al_left,   None),
            (row.tipo,         al_center, None),
            (valor,            al_right,  FMT_MOEDA),
            (saldo,            al_right,  FMT_MOEDA),
            (row.banco_origem, al_left,   None),
            (row.forma_pagamento, al_left, None),
            (str(row.categoria) if pd.notna(row.categoria) else "", al_left, None),
            (str(row.apuracao)  if pd.notna(row.apuracao)  else "", al_left, None),
            (str(row.observacao) if pd.notna(row.observacao) else "", al_left, None),
        ]

        for col_idx, (value, align, fmt) in enumerate(linha, start=1):
            _cell(ws, row_idx, col_idx, value, fill, font_data, align, borda, fmt)

    # ── Linha de totais ───────────────────────────────────────────────────────
    total_row = len(df) + 2

    entradas = df[df["tipo"] == "Entrada"]["valor"].sum()
    saidas   = df[df["tipo"] == "Saída"]["valor"].sum()
    saldo_liq = entradas - saidas

    totais = [
        ("TOTAL ENTRADAS", entradas, f"{len(df[df['tipo']=='Entrada'])} lançamentos"),
        ("TOTAL SAÍDAS",   saidas,   f"{len(df[df['tipo']=='Saída'])} lançamentos"),
        ("SALDO DO PERÍODO", saldo_liq, "Entradas − Saídas"),
    ]

    for i, (label, valor, obs) in enumerate(totais):
        r = total_row + i
        ws.row_dimensions[r].height = 20
        for col in range(1, len(COLUNAS) + 1):
            _cell(ws, r, col, None, fill_total, font_total, al_center, borda)
        _cell(ws, r, 2, label,  fill_total, font_total, al_right,  borda)
        _cell(ws, r, 4, valor,  fill_total, font_total, al_right,  borda, FMT_MOEDA)
        _cell(ws, r, 10, obs,   fill_total, font_total, al_left,   borda)

    # ── Filtro, freeze e range ────────────────────────────────────────────────
    last_row = total_row + len(totais) - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{len(df) + 1}"
    ws.freeze_panes    = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
